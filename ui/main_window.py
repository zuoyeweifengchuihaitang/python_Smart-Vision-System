# -*- coding: utf-8 -*-
import time
import cv2
import os
import numpy as np
import pandas as pd
from PIL import Image, ImageDraw, ImageFont

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.pyplot as plt

from core.models import yolo_person
from database.operations import startup_self_check, register_face
from ui.widgets import ClickLabel
from ui.dialogs import CaptureWindow, ManageDialog
from ui.worker import VisionEngine
from config import FONT_PATH, LOG_PATH

class SmartVisionApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(u"AI 综合管理系统 v18.0 (身份可视版)")
        self.resize(1300, 850)
        
        self.f_db, self.bl, self.wl = startup_self_check()
        self.engine = None
        self.line_step, self.pts, self.curr_video = 0, [], None
        self.roi_step, self.roi_pts = 0, []
        self.temp_dims = (640, 480)
        self.current_density_count = 0
        
        self.init_ui()
        self.apply_style()

    def init_ui(self):
        c = QWidget()
        self.setCentralWidget(c)
        l = QHBoxLayout(c)
        
        nav = QVBoxLayout()
        btns = [
            (u"🎥 实时监控", self.act_face), (u"🖼️ 图片识别", self.act_img), (u"🎬 视频分析", self.act_video), 
            (u"🚶 视频流量统计", self.act_flow), (u"👥 人群密度统计", self.act_density), 
            (u"📸 摄像头录入", self.act_reg_cam), (u"📂 文件导入人像", self.act_reg_file),
            (u"⚙️ 人脸管理", self.act_manage), (u"📊 数据看板", self.act_dash), 
            (u"📥 导出 Excel", self.act_excel), (u"⏹️ 停止任务", self.stop)
        ]
        
        for t, f in btns:
            b = QPushButton(t)
            b.clicked.connect(f)
            nav.addWidget(b)
            
        nav.addStretch()
        self.info = QLabel(u"等待指令")
        nav.addWidget(self.info)
        l.addLayout(nav, 1)

        self.view = ClickLabel(u"视频流显示区域")
        self.view.setAlignment(Qt.AlignCenter)
        self.view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.view.clicked_pos.connect(self.on_view_click)
        l.addWidget(self.view, 4)
        
        side = QVBoxLayout()
        side.addWidget(QLabel(u"实时日志:"))
        self.log_p = QListWidget()
        self.log_p.setStyleSheet("background:#1a252f;color:#2ecc71;")
        l.addLayout(side, 1)
        side.addWidget(self.log_p)

    def apply_style(self):
        self.setStyleSheet("QMainWindow{background:#2c3e50;} QPushButton{background:#34495e;color:white;padding:12px;} QLabel{color:white;}")

    def act_face(self): 
        self.stop()
        self.engine = VisionEngine(0, self.f_db, self.bl, self.wl, 'face')
        self._connect_engine()
        self.engine.start()
    
    def act_img(self):
        self.stop()
        p, _ = QFileDialog.getOpenFileName(self, u"选图", "", "Img (*.jpg *.png)")
        if p:
            img = cv2.imdecode(np.fromfile(p, dtype=np.uint8), cv2.IMREAD_COLOR)
            self.engine = VisionEngine(img, self.f_db, self.bl, self.wl, 'face')
            self._connect_engine()
            self.engine.start()

    def act_video(self):
        self.stop()
        p, _ = QFileDialog.getOpenFileName(self, u"选视频", "", "Video (*.mp4 *.avi)")
        if p:
            self.engine = VisionEngine(p, self.f_db, self.bl, self.wl, 'face')
            self._connect_engine()
            self.engine.start()

    def act_flow(self):
        self.stop()
        p, _ = QFileDialog.getOpenFileName(self, u"流量视频")
        if not p:
            return
        self.curr_video = p
        cap = cv2.VideoCapture(p)
        ret, frame = cap.read()
        cap.release()
        
        if ret:
            self.temp_dims = (frame.shape[1], frame.shape[0])
            self.upd(frame)
            self.line_step = 1
            self.pts = []
            QMessageBox.information(self, u"配置", u"请在画面中点击【起点 A】")
            self.info.setText(u"请点击画面确定起点 A")

    def act_density(self):
        self.stop()
        p, _ = QFileDialog.getOpenFileName(self, u"人群密度视频", "", "Video (*.mp4 *.avi)")
        if not p:
            return
        self.curr_video = p
        cap = cv2.VideoCapture(p)
        ret, frame = cap.read()
        cap.release()
        
        if ret:
            self.temp_dims = (frame.shape[1], frame.shape[0])
            self.upd(frame)
            self.roi_step = 1
            self.roi_pts = []
            QMessageBox.information(self, u"ROI配置", u"请点击画面确定ROI【左上角】，然后【右下角】。")
            self.info.setText(u"步骤1: 点击左上角")

    def _connect_engine(self):
        if self.engine:
            self.engine.frame_ready.connect(self.upd)
            self.engine.flow_ready.connect(self.upd_f)
            self.engine.log_signal.connect(self.push)
            self.engine.count_ready.connect(self.handle_density_count)

    def handle_density_count(self, count):
        self.current_density_count = count

    def get_real_coords(self, click_x, click_y):
        lbl_w, lbl_h = self.view.width(), self.view.height()
        img_w, img_h = self.temp_dims
        if img_w == 0 or img_h == 0:
            return 0, 0
        ratio_img = img_w / img_h
        ratio_lbl = lbl_w / lbl_h
        if ratio_lbl > ratio_img:
            drawn_h = lbl_h
            drawn_w = lbl_h * ratio_img
            dx = (lbl_w - drawn_w) / 2
            dy = 0
        else:
            drawn_w = lbl_w
            drawn_h = lbl_w / ratio_img
            dx = 0
            dy = (lbl_h - drawn_h) / 2
        return int((click_x - dx) * (img_w / drawn_w)), int((click_y - dy) * (img_h / drawn_h))

    def on_view_click(self, x, y):
        if self.line_step > 0:
            rx, ry = self.get_real_coords(x, y)
            if self.line_step == 1:
                self.pts.append((rx, ry))
                self.line_step = 2
                self.info.setText(u"请点击【终点 B】")
            elif self.line_step == 2:
                self.pts.append((rx, ry))
                self.line_step = 3
                self.info.setText(u"请点击【进入(IN)】的一侧")
            elif self.line_step == 3:
                (x1, y1), (x2, y2) = self.pts[0], self.pts[1]
                sign = 1 if (x2 - x1) * (ry - y1) - (y2 - y1) * (rx - x1) > 0 else -1
                config = {'p1': self.pts[0], 'p2': self.pts[1], 'sign': sign}
                self.engine = VisionEngine(self.curr_video, mode='flow', flow_config=config)
                self._connect_engine()
                self.engine.start()
                self.line_step = 0
                self.info.setText(u"流量统计运行中...")
        elif self.roi_step > 0:
            rx, ry = self.get_real_coords(x, y)
            if self.roi_step == 1:
                self.roi_pts.append((rx, ry))
                self.roi_step = 2
                self.info.setText(u"步骤2: 点击右下角")
            elif self.roi_step == 2:
                self.roi_pts.append((rx, ry))
                x1 = min(self.roi_pts[0][0], self.roi_pts[1][0])
                y1 = min(self.roi_pts[0][1], self.roi_pts[1][1])
                x2 = max(self.roi_pts[0][0], self.roi_pts[1][0])
                y2 = max(self.roi_pts[0][1], self.roi_pts[1][1])
                
                if x2 - x1 < 50 or y2 - y1 < 50:
                    QMessageBox.warning(self, u"无效ROI", u"ROI区域太小，请重新绘制。")
                    self.roi_step = 0
                    return

                threshold, ok = QInputDialog.getInt(self, u"密度阈值", u"告警阈值（间隔内最大人数）：", 20, 1, 1000)
                if not ok:
                    self.roi_step = 0
                    return
                interval, ok = QInputDialog.getInt(self, u"告警间隔", u"检查间隔（秒）：", 10, 1, 600)
                if not ok:
                    self.roi_step = 0
                    return

                config = {'roi': (x1, y1, x2, y2), 'threshold': threshold, 'alert_interval': interval}
                self.engine = VisionEngine(self.curr_video, mode='density', density_config=config)
                self._connect_engine()
                self.engine.start()
                self.roi_step = 0
                self.info.setText(u"人群密度统计运行中...")

    def upd(self, d):
        if self.engine and self.engine.mode == 'density':
            roi = self.engine.density_config.get('roi')
            count = self.current_density_count

            if roi:
                rx1, ry1, rx2, ry2 = roi
                cv2.rectangle(d, (rx1, ry1), (rx2, ry2), (0, 255, 0), 6)

            if count > self.engine.density_threshold:
                heat = np.zeros(d.shape[:2], dtype=np.float32)
                res = yolo_person(d, classes=[0], verbose=False, conf=0.3)[0]
                boxes = res.boxes.xyxy.cpu().numpy().astype(int)
                for box in boxes:
                    cx = (box[0] + box[2]) // 2
                    cy = (box[1] + box[3]) // 2
                    if roi and rx1 <= cx <= rx2 and ry1 <= cy <= ry2:
                        cv2.circle(heat, (cx, cy), 60, 1.5, -1)

                heat = cv2.GaussianBlur(heat, (0, 0), sigmaX=30)
                heat_norm = cv2.normalize(heat, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
                heat_color = cv2.applyColorMap(heat_norm, cv2.COLORMAP_JET)
                d = cv2.addWeighted(d, 0.6, heat_color, 0.4, 0)
                cv2.rectangle(d, (0, 0), (d.shape[1], d.shape[0]), (0, 0, 255), 15)

            pil_img = Image.fromarray(cv2.cvtColor(d, cv2.COLOR_BGR2RGB))
            draw = ImageDraw.Draw(pil_img)
            try:
                font_pil = ImageFont.truetype(FONT_PATH, 100)
            except:
                font_pil = ImageFont.load_default()

            text = f"实时密度: {count} 人"
            fill_color = (255, 0, 0) if count > self.engine.density_threshold else (0, 255, 0)
            draw.text((50, 50), text, font=font_pil, fill=fill_color)
            d = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)

        self.temp_dims = (d.shape[1], d.shape[0])
        qt_img = QImage(d.data, d.shape[1], d.shape[0], d.shape[1]*3, QImage.Format_RGB888).rgbSwapped()
        pix = QPixmap.fromImage(qt_img)
        scaled_pix = pix.scaled(self.view.contentsRect().size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.view.setPixmap(scaled_pix)

    def upd_f(self, s): 
        self.info.setText(f"IN:{s['in']} | OUT:{s['out']} | {s['elapsed']}s")

    def push(self, src, n, s): 
        it = QListWidgetItem(f"[{time.strftime('%H:%M:%S')}] [{src}] {n} - {s}")
        it.setForeground(QColor("#2ecc71" if "白名单" in s or "进入" in s else "#e74c3c"))
        self.log_p.insertItem(0, it)

    def act_reg_cam(self):
        n, ok = QInputDialog.getText(self, u"录入", u"输入姓名:")
        if ok and n:
            dlg = CaptureWindow(self)
            if dlg.exec_() == QDialog.Accepted:
                self.do_reg(dlg.captured_frame, n)

    def act_reg_file(self):
        n, ok = QInputDialog.getText(self, u"录入", u"输入姓名:")
        if ok and n:
            p, _ = QFileDialog.getOpenFileName(self, u"选图", "", "Img (*.jpg *.png)")
            if p:
                self.do_reg(cv2.imdecode(np.fromfile(p, dtype=np.uint8), cv2.IMREAD_COLOR), n)

    def do_reg(self, img, n):
        if img is None:
            return
        cat, ok = QInputDialog.getItem(self, u"分类", u"类型:", [u"白名单", u"黑名单"], 0, False)
        if ok:
            succ, msg = register_face(img, n, '1' if u"黑" in cat else '2')
            if succ:
                self.f_db, self.bl, self.wl = startup_self_check()
                self.push(u"录入", n, u"成功")
                QMessageBox.information(self, u"成功", msg)
            else:
                QMessageBox.warning(self, u"失败", msg)

    def act_manage(self):
        dlg = ManageDialog(self.f_db, self.bl, self.wl, self)
        dlg.exec_() 
        print("🔄 管理界面已关闭，正在刷新数据...")
        
        self.f_db, self.bl, self.wl = startup_self_check()

        if self.engine:
            self.engine.face_db = self.f_db
            self.engine.bl = self.bl
            self.engine.wl = self.wl
            print(f"🚀 摄像头数据已更新：当前库中剩余 {len(self.f_db)} 人")

    def act_dash(self):
        if not os.path.exists(LOG_PATH):
            QMessageBox.warning(self, u"空", u"暂无记录")
            return

        d = QDialog(self)
        d.setWindowTitle(u"数据分布")
        d.resize(700, 500)
        v = QVBoxLayout(d)

        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS']
        plt.rcParams['axes.unicode_minus'] = False

        df = pd.read_csv(LOG_PATH, encoding='utf-8-sig')

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 5))
        canvas = FigureCanvas(fig)
        v.addWidget(canvas)

        df[u'状态'].value_counts().plot(kind='pie', ax=ax1, autopct='%1.1f%%')
        ax1.set_title(u"状态分布")

        df[u'来源'].value_counts().plot(kind='bar', ax=ax2)
        ax2.set_title(u"来源统计")
        plt.setp(ax2.get_xticklabels(), rotation=45, ha='right')

        d.exec_()

    def act_excel(self):
        if not os.path.exists(LOG_PATH):
            QMessageBox.warning(self, u"提示", u"暂无数据")
            return
        p, _ = QFileDialog.getSaveFileName(self, u"保存", "Report.xlsx", "Excel (*.xlsx)")
        if p:
            try:
                df = pd.read_csv(LOG_PATH, encoding='utf-8-sig')
                with pd.ExcelWriter(p, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Log')
                    from openpyxl.styles import PatternFill
                    ws = writer.sheets['Log']
                    red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    for r in range(2, ws.max_row+1):
                        if u"黑名单" in str(ws.cell(r, 4).value):
                            for c in range(1, 6):
                                ws.cell(r, c).fill = red
                QMessageBox.information(self, u"成功", u"导出完成")
            except Exception as e:
                QMessageBox.critical(self, u"错误", str(e))

    def stop(self):
        self.line_step = 0
        self.roi_step = 0
        if self.engine:
            self.engine.stop()
            self.engine = None
        self.view.clear()
        self.info.setText(u"已停止")